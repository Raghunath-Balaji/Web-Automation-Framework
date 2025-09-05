import os
import logging
import configparser  # for loading .properties files
import random  # for Java Random
import threading  # for ThreadLocal analogues
from selenium import webdriver  # Selenium WebDriver
from selenium.webdriver.chrome.service import Service  # ChromeDriver service
from selenium.webdriver.chrome.options import Options  # Chrome options
from selenium.webdriver.firefox.options import Options as FirefoxOptions  # Firefox options
from selenium.webdriver.support.ui import WebDriverWait  # Equivalent to Selenium WebDriverWait
from openpyxl import load_workbook

# Incompatibility notes:
# java.io.FileInputStream -> Python uses open()
# java.net.URL -> use urllib.parse if needed
# java.text.SimpleDateFormat -> use datetime.strftime; SimpleDateFormat stub below
# com.google.common.collect.ImmutableMap -> use dict literals
# io.appium.java_client.* -> Requires Appium-Python-Client; stub methods if needed
# com.codoid.products.fillo.* -> Fillo not in Python; stub methods where required
# org.apache.poi.* -> use openpyxl or pandas; stub methods where required
# from utilities.ExcelReader import ExcelReader  # ExcelReader port stub

# Configure module logger
logger = logging.getLogger(__name__)


class DriverEngine:
    # --- Chunk 1: Configuration Fields ---
    property: configparser.ConfigParser = None
    brand: str = None
    environment_name: str = None
    build_type: str = None
    browser_name: str = None
    platform_name: str = None
    headless_browser: str = None
    field_validation: str = None
    web_app_url: str = None
    op_system: str = None
    browserstack_userid: str = None
    browserstack_access_key: str = None
    execution_time: str = None
    device_name: str = None
    device_os_version: str = None
    # reader: ExcelReader = ExcelReader()
    config_sheet_path: str = os.path.abspath("resources/test_data/configuration.xlsx")

    # --- Chunk 2: Execution & Reporting Fields ---
    element_wait_in_seconds: int = 60
    wait: WebDriverWait = None
    user_device_input: str = None
    simpleformat = None  # SimpleDateFormat stub; use datetime.strftime where needed
    test_data_type: str = None
    sheet_path: str = ""
    storage_connection_string: str = None
    os_name: str = None
    global_row_number: int = 0
    performance_time_update: str = None
    window_heading: str = None
    report_sheet_name: str = None
    report_sheet_path: str = None
    report_row: int = 2
    last_used_row: int = 2
    sheet_exists_flag: bool = False
    rng: random.Random = random.Random()

    # --- Chunk 3: ThreadLocals & Misc Fields ---
    str_value: str = None  # corresponds to Java 'str'
    location_change_start_order_tf: bool = False
    tl_scenario_name: threading.local = threading.local()
    tl_driver: threading.local = threading.local()
    multi_capabilities: threading.local = threading.local()  # MutableCapabilities stub
    chrome_options_local: threading.local = threading.local()
    firefox_options_local: threading.local = threading.local()
    web_driver_wait: threading.local = threading.local()
    testrecord_set = None  # Recordset stub - requires Python equivalent (Fillo)
    tlcon_test_data = None  # Connection stub - requires DB or file connection
    ui_automator2_options = None  # UiAutomator2Options stub - requires Appium
    member_driver1: webdriver.Remote = None  # RemoteWebDriver stub
    fillo = None  # stub for Fillo library not available in Python
    con_config = None  # Connection stub
    build_name: str = None

    def load_properties(self, path: str) -> configparser.ConfigParser:
        """
        Load a Java .properties file and return a ConfigParser.
        """
        config = configparser.ConfigParser()
        config.read(path)
        return config

    def load_inputs(self) -> None:
        global property, brand, environmentName, buildType, browserName, \
            platformName, headlessBrowser, fieldValidation, webAppUrl, \
            getbrowserstackuserid, getbrowserstackacesskey, opSystem

        # Load properties file (if it's a separate file for browser-related configs)
        # The original Java code has property = loadProperty(BROWSER_PATH);
        # If BROWSER_PATH leads to a .properties file, use configparser.
        # If it's directly related to the 'Config' sheet data, this line might be redundant
        # or serves a different purpose. I'll assume it's for external properties.
        #  property = self.load_properties(BROWSER_PATH)

        # In Python, you'd typically handle exceptions with try...except
        try:
            rsConfig = self.con_config.executeQuery("SELECT * FROM Config")
            if rsConfig.next():  # Move to the first record
                brand = rsConfig.getField("Brand")
                environmentName = rsConfig.getField("Env")
                buildType = rsConfig.getField("BuildType")
                browserName = rsConfig.getField("Browser")
                platformName = rsConfig.getField("Platform")
                headlessBrowser = rsConfig.getField("HeadlessBrowser")
                fieldValidation = rsConfig.getField("FieldValidation")
                # deviceName = rsConfig.getField("DeviceName")
                # deviceOSVersion = rsConfig.getField("DeviceOSVersion")

            print("Brand :", brand)
            print("Environment_Name :", environmentName)
            print("Browser_Name :", browserName)
            print("Platform_Name :", platformName)
            print("Headless_Browser :", headlessBrowser)

            if environmentName and environmentName.lower() == "live":
                webAppUrl = f"https://{brand}.com"
            elif environmentName and environmentName.lower() == "beta":
                if brand and brand.lower() == "mca":
                    webAppUrl = "https://beta.mcalistersdeli.com"
                else:
                    webAppUrl = f"https://beta.{brand}.com"
            elif brand and brand.lower() == "schlotzskys":
                webAppUrl = f"https://{environmentName}.{brand}.com"
            else:
                webAppUrl = f"https://{brand}.{environmentName}.focusbrands.com"

            getbrowserstackuserid = rsConfig.getField("BrowserStackUserID")
            getbrowserstackacesskey = rsConfig.getField("BrowserStackAcessKey")
            opSystem = rsConfig.getField("OpreatingSystem")

            rsConfig.close()

        except Exception as e:
            # Catching a general exception. Be more specific if possible.
            print(f"An error occurred during load_inputs: {e}")
            # Handle specific exceptions like InvalidFormatException, IOException, FilloException
            # For FilloException, you'd catch the specific exception from your data access library.
            # For IO exceptions, use OSError or more specific file-related exceptions.

    def launch_browser(self) -> None:
        """
        Launch browser using initialized driver and navigate to web_app_url.
        """
        self.initialize_driver()
        driver = self.get_driver()
        if self.web_app_url and driver:
            driver.get(self.web_app_url)

    def is_local(self) -> bool:
        """
        Determine if execution is local (not BrowserStack).
        """
        if self.platform_name and len(self.platform_name) >= 3:
            return not self.platform_name[:3].lower() == "bs_"
        return True

    def get_driver(self) -> webdriver.Remote:
        """
        Retrieve the current thread-local driver instance.
        """
        return getattr(self.tl_driver, "driver", None)

    def get_email_from_brand_sheet(
            file_path: str,
            sheet_name: str,
            type_value: str
    ) -> dict[str, str]:

        # --- 1) Open workbook & grab the sheet ---
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # --- 2) Read headers and find column indexes ---
        headers = [cell.value for cell in ws[1]]  # first row = header
        # map header name -> 1-based column index
        idx = {name: i + 1 for i, name in enumerate(headers)}

        # shorthand lookups
        TYPE_COL = idx['Type']
        FLAG_COL = idx['Flag']
        EMAIL_COL = idx['Email']
        ROWNUM_COL = idx.get('RowNum')  # assume there’s a RowNum column

        # --- 3) Collect all candidate rows ---
        candidates = []
        for row in ws.iter_rows(min_row=2):  # skip header
            if row[TYPE_COL - 1].value == type_value and row[FLAG_COL - 1].value == 'No':
                candidates.append(row)

        if not candidates:
            raise ValueError(f"No un-flagged rows for Type='{type_value}'")

        chosen = random.choice(candidates)

        result = {headers[i]: chosen[i].value for i in range(len(headers))}

        chosen[FLAG_COL - 1].value = 'Yes'
        wb.save(file_path)

        return result

    def get_result_from_brand_sheet(file_path: str, sheet_name: str, scenario_name: str) -> dict[str, str]:

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
        except Exception as e:
            logging.error(f"Couldn’t open sheet '{sheet_name}' in '{file_path}': {e}")
            return {}

        # 1) Read header row
        headers = [cell.value for cell in ws[1]]
        if "ScenarioName" not in headers:
            logging.error(f"'ScenarioName' column not found in {sheet_name}")
            return {}
        scenario_idx = headers.index("ScenarioName")

        # 2) Scan data rows for the first match
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[scenario_idx] == scenario_name:
                # 3) Build and return a dict of column→value
                return {headers[i]: row[i] for i in range(len(headers))}

        # 4) No match found
        logging.warning(f"No row with ScenarioName='{scenario_name}' in {sheet_name}")
        return {}

    def get_chrome_driver(self) -> webdriver.Remote:
        """
        Equivalent of getChromeDriver(): init Chrome with options.
        """
        options = Options()
        options.add_argument("--disable-notifications")
        if self.headless_browser and self.headless_browser.lower() == "true":
            options.add_argument("--headless")
        driver = webdriver.Chrome(service=Service(), options=options)
        setattr(self.tl_driver, 'driver', driver)
        return driver

    def get_firefox_driver(self) -> webdriver.Remote:
        """
        Equivalent of getFireFoxDriver(): init Firefox with options and profile.
        """
        options = FirefoxOptions()
        options.add_argument("--disable-notifications")
        # FirefoxProfile not directly supported; use capabilities if needed
        if self.headless_browser and self.headless_browser.lower() == "true":
            options.add_argument("--headless")
        driver = webdriver.Firefox(options=options)
        setattr(self.tl_driver, 'driver', driver)
        return driver

    def get_safari_driver(self) -> webdriver.Remote:
        """
        Equivalent of getSafariDriver(): init Safari.
        """
        try:
            driver = webdriver.Safari()
        except Exception as e:
            raise NotImplementedError("Safari WebDriver not supported on this platform: " + str(e))
        setattr(self.tl_driver, 'driver', driver)
        return driver

    def get_android_chrome_browser(self) -> webdriver.Remote:
        """
        Equivalent of getAndroidChromeBrowser(): requires Appium Python client.
        """
        raise NotImplementedError("get_android_chrome_browser requires Appium client and mobile context")

    def get_ios_safari_browser(self) -> webdriver.Remote:
        """
        Equivalent of getIOSSafariBrowser(): requires Appium Python client and iOS context.
        """
        raise NotImplementedError("get_ios_safari_browser requires Appium client and iOS context")

    def get_remote_desktop_driver(self) -> webdriver.Remote:

        self.multi_capabilities.set({})

        browserstack_options = {}

        if opSystem.lower() == "windows":
            browserstack_options["os"] = "Windows"
            browserstack_options["osVersion"] = "11"
        elif opSystem.lower() == "mac":
            browserstack_options["os"] = "OS X"
            browserstack_options["osVersion"] = "Sonoma"

        options = Options()
        options.add_argument("--disable-notifications")

        browserstack_options["browserName"] = browserName
        browserstack_options["browserVersion"] = "latest"
        browserstack_options["local"] = False
        browserstack_options["geoLocation"] = "US-GA"
        browserstack_options["networkLogs"] = True
        browserstack_options["selfHeal"] = True
        browserstack_options["sessionName"] = (
            f"{self.tl_scenario_name.get()} - Browser : {browserName} OS: {opSystem}"
        )
        browserstack_options["buildName"] = self.build_name

        self.multi_capabilities.get()["bstack:options"] = browserstack_options

        final_capabilities = options.to_capabilities()
        final_capabilities["bstack:options"] = browserstack_options

        browserstack_url = (
            f"http://{getbrowserstackuserid}:{getbrowserstackacesskey}"
            "@hub-cloud.browserstack.com/wd/hub"
        )

        return webdriver.Remote(command_executor=browserstack_url, options=self.multi_capabilities.get())

        # raise NotImplementedError("get_remote_desktop_driver requires remote grid configuration")

    def get_remote_desktop_driver_with_url(self, grid_url: str) -> webdriver.Remote:
        """
        Equivalent of getRemoteDesktopDriver(String): requires remote grid URL.
        """
        raise NotImplementedError("get_remote_desktop_driver_with_url requires remote grid URL")

    def get_remote_android_mobile_driver(self) -> webdriver.Remote:
        """
        Equivalent of getRemoteAndroidMobileDriver(): requires Appium Python client and Android context.
        """
        raise NotImplementedError("get_remote_android_mobile_driver requires Appium client and Android context")

    def get_remote_ios_mobile_driver(self) -> webdriver.Remote:
        """
        Equivalent of getRemoteIOSMobileDriver(): requires Appium Python client and iOS context.
        """
        raise NotImplementedError("get_remote_ios_mobile_driver requires Appium client and iOS context")

    def get_bs_ios_safari_browser(self) -> webdriver.Remote:
        """
        Equivalent of getBSIOSafariBrowser(): requires BrowserStack setup and credentials.
        """
        raise NotImplementedError("get_bs_ios_safari_browser requires BrowserStack setup and credentials")

    def initialize_driver(self) -> None:
        driver = None
        if self.is_local():
            browser_name_upper = self.browser_name.upper()
            if browser_name_upper == "CHROME":
                driver = self.get_chrome_driver()
            elif browser_name_upper == "SAFARI":
                driver = self.get_safari_driver()
            elif browser_name_upper == "FIREFOX":
                driver = self.get_firefox_driver()
        else:
            platform_name_upper = self.platform_name.upper()
            if platform_name_upper == "BS_WEB":
                driver = self.get_remote_desktop_driver()
            elif platform_name_upper == "BS_MOBILEWEB":
                if self.browser_name.lower() == "bs_mobilechrome":
                    driver = self.get_remote_android_mobile_driver()
                elif self.browser_name.lower() == "mobilechrome":
                    driver = self.get_android_chrome_browser()
                elif self.browser_name.lower() == "mobilesafari":
                    driver = self.get_ios_safari_browser()
                else:
                    driver = self.get_remote_ios_mobile_driver()

        self.tl_driver.set(driver)
        self.web_driver_wait.set(WebDriverWait(self.tl_driver.get(), self.element_wait_in_seconds))
        self.launch_browser()
