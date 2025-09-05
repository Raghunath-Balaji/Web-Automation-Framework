import os
from typing import Optional, List, Dict
import logging
from utilities.Utilities import Utilities
from openpyxl import *
from openpyxl.worksheet.worksheet import Worksheet
from collections import OrderedDict
from openpyxl.cell.cell import Cell as openpyxl_Cell
from io import BytesIO
from azure.storage.blob import BlobServiceClient
from openpyxl.utils import get_column_letter



class ExcelReader(Utilities):

    def get_data_by_name(self, excel_file_path: str, sheet_name: str) -> List[Dict[str, str]]:

        sheet = self.get_sheet_by_name(excel_file_path, sheet_name)
        if sheet is None:
            return []
        return self.read_sheet(sheet)

    def get_data_by_index( self, excel_file_path: str, sheet_index: int ) -> List[Dict[str, str]]:

        try:
            workbook = load_workbook(excel_file_path)
        except Exception as e:
            return []

        sheet = self.get_sheet_by_index(workbook, sheet_index)
        if sheet is None:
            return []

        return self.read_sheet(sheet)

    def get_work_book(self, excel_file_path: str):
        try:
            workbook = load_workbook(excel_file_path)
            return workbook
        except Exception as e:
            return None

    def get_sheet_by_index(self, workbook, sheet_index: int):
        try:
            sheet = workbook.worksheets[sheet_index]
            return sheet
        except IndexError:
            return None

    def get_sheet_by_name(self, excel_file_path: str, sheet_name: str) -> Optional[Worksheet]:
        wb = None

        try:
            wb = self.get_work_book(excel_file_path)

            return wb[sheet_name]

        except KeyError:
            logging.error("Sheet '%s' not found in '%s'", sheet_name, excel_file_path)
        except Exception:
            logging.exception("Failed to load sheet '%s' from '%s'", sheet_name, excel_file_path)

        return None

    def get_header_row_number(self, sheet):

        max_row = sheet.max_row
        max_col = sheet.max_column
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    return row - 1
        return -1

    def read_sheet(self, sheet: Worksheet) -> List[Dict[str, str]]:

        header_idx = self.get_header_row_number(sheet)
        if header_idx < 0:
            return []  # no header ⇒ no data

        # openpyxl is 1-based
        header_row_num = header_idx + 1
        total_columns = sheet.max_column

        # read header names
        headers = []
        for col in range(1, total_columns + 1):
            raw = sheet.cell(row=header_row_num, column=col).value
            headers.append(str(raw) if raw is not None else f"Column{col}")

        rows: List[Dict[str, str]] = []
        # data starts one row after header
        for r in range(header_row_num + 1, sheet.max_row + 1):
            row_map = OrderedDict()
            for col_idx, col_name in enumerate(headers, start=1):
                val = sheet.cell(row=r, column=col_idx).value
                # coerce to string so Python dict[str,str]
                row_map[col_name] = str(val) if val is not None else ""
            rows.append(row_map)

        return rows

    def get_row(self, sheet, row_number: int):

        excel_row = row_number + 1

        if excel_row < 1:
            return None
        rows = list(sheet.iter_rows(min_row=1, max_row=excel_row, values_only=False))
        return list(rows[0]) if rows else None

    def get_cell_value(self, sheet: Worksheet, row_cells: List[openpyxl_Cell], current_column: int) -> Dict[str, str]:

        column_map: Dict[str, str] = {}

        header_idx = self.get_header_row_number(sheet)
        if header_idx < 0:
            return column_map
        header_excel_row = header_idx + 1
        header_cell = sheet.cell(
            row=header_excel_row,
            column=current_column + 1
        )

        if header_cell.value is None:
            return column_map
        header_name = str(header_cell.value)

        if row_cells is None or current_column >= len(row_cells):
            column_map[header_name] = ""
            return column_map

        cell: openpyxl_Cell = row_cells[current_column]
        val = cell.value

        if val is None:
            column_map[header_name] = ""
        else:
            column_map[header_name] = str(val)

        return column_map

    # def get_excel_cell_data(self, excel_file_path: str, sheet_identifier: str | int, cell_reference: str):
    #     workbook = self.get_work_book(excel_file_path)
    #
    #     if not workbook:
    #         print(f"Error: Could not load workbook from '{excel_file_path}'")
    #         return None
    #
    #     sheet = None
    #     if isinstance(sheet_identifier, str):
    #         sheet = self.get_sheet_by_name(excel_file_path, sheet_identifier)
    #     elif isinstance(sheet_identifier, int):
    #         sheet = self.get_sheet_by_index(workbook, sheet_identifier)
    #     else:
    #         print("Error: sheet_identifier must be a string (sheet name) or an integer (sheet index).")
    #         return None
    #
    #     if not sheet:
    #         print(f"Error: Sheet '{sheet_identifier}' not found in '{excel_file_path}'")
    #         return None
    #
    #     try:
    #         row, col = openpyxl_Cell.coordinate_to_tuple(cell_reference)
    #         cell_value = sheet.cell(row=row, column=col).value
    #         return cell_value
    #     except Exception as e:
    #         print(f"Error accessing cell '{cell_reference}': {e}")
    #         return None

    def get_data_as_string(self, FILE_PATH, SHEET_IDENTIFIER, HEADER_SUBSTRING):

        workbook = self.get_work_book(FILE_PATH)
        if not workbook:
            print(f"Error: Could not load workbook from '{FILE_PATH}'")
        else:
            sheet = None
            if type(SHEET_IDENTIFIER) is str:

                print("type of sheet is string")
                sheet = self.get_sheet_by_name(FILE_PATH, SHEET_IDENTIFIER)

            elif type(SHEET_IDENTIFIER) is int:

                sheet = self.get_sheet_by_index(workbook, SHEET_IDENTIFIER)

            else:
                print("Error: SHEET_IDENTIFIER must be a string (sheet name) or an integer (sheet index).")

            if not sheet:
                print(f"Error: Sheet '{SHEET_IDENTIFIER}' not found in '{FILE_PATH}'")
            else:
                all_data_rows = self.read_sheet(sheet)

                if not all_data_rows:
                    print(f"No data found in sheet '{SHEET_IDENTIFIER}'.")
                else:
                    first_data_row = all_data_rows[0]

                    found_header_key = None
                    for header_key in first_data_row.keys():
                        if HEADER_SUBSTRING.lower() in header_key.lower():
                            found_header_key = header_key
                            break

                    if found_header_key:
                        name_data = first_data_row[found_header_key]
                        return name_data
                    else:
                        return "No data found matching the header substring."


def read_azure_data(excel_name: str, sheet_name: str) -> Optional[Worksheet]:
    try:

        conn_str = os.getenv("APPPACKAGE")
        if not conn_str:
            raise RuntimeError("Missing APPPACKAGE environment variable")

        client = BlobServiceClient.from_connection_string(conn_str)
        container = client.get_container_client("moes-mobile-testdata")
        blob = container.get_blob_client(excel_name)

        stream = blob.download_blob()
        data = stream.readall()

        wb = load_workbook(filename=BytesIO(data), data_only=True)
        return wb[sheet_name]

    except Exception:
        logging.exception(
            "Failed to read Azure Excel '%s' sheet '%s'",
            excel_name, sheet_name
        )
        return None


def set_cell_data(
        sheet_path: str,
        sheet_name: str,
        col_name: str,
        row_num: int,
        data: str
) -> bool:
    """
    Write `data` into (sheet_name, col_name, row_num) of the workbook at sheet_path.
    Returns True on success, False on any failure (missing sheet/column, IO error, etc.).
    """
    try:
        wb = load_workbook(sheet_path)
        sheet = wb[sheet_name]  # KeyError if sheet_name not found

        # 1) find the column index by matching header row (row 1)
        header_row = list(sheet[1])
        col_idx = next(
            (i + 1 for i, cell in enumerate(header_row)
             if str(cell.value or "").strip().lower() == col_name.lower()),
            None
        )
        if col_idx is None:
            return False

        # 2) set the cell (openpyxl auto-creates row/cell if needed)
        cell = sheet.cell(row=row_num, column=col_idx)
        cell.value = data

        # 3) auto-size the column
        letter = get_column_letter(col_idx)
        max_length = max(
            (len(str(c.value)) for c in sheet[letter] if c.value is not None),
            default=len(col_name)
        )
        sheet.column_dimensions[letter].width = max_length + 2

        # 4) save changes
        wb.save(sheet_path)
        wb.close()
        return True

    except Exception:
        logging.exception(
            "Failed to set cell data in %r [%s!%s@%d]",
            sheet_path, sheet_name, col_name, row_num
        )
        return False


def set_report_cell_data(
        report_sheet_path: str,
        sheet_name: str,
        col_name: str,
        row_num: int,
        data: str
) -> bool:
    """
    Exactly the same as set_cell_data, but writes to report_sheet_path instead.
    """
    try:
        wb = load_workbook(report_sheet_path)
        sheet = wb[sheet_name]

        header_row = list(sheet[1])
        col_idx = next(
            (i + 1 for i, cell in enumerate(header_row)
             if str(cell.value or "").strip().lower() == col_name.lower()),
            None
        )
        if col_idx is None:
            return False

        cell = sheet.cell(row=row_num, column=col_idx)
        cell.value = data

        letter = get_column_letter(col_idx)
        max_length = max(
            (len(str(c.value)) for c in sheet[letter] if c.value is not None),
            default=len(col_name)
        )
        sheet.column_dimensions[letter].width = max_length + 2

        wb.save(report_sheet_path)
        wb.close()
        return True

    except Exception:
        logging.exception(
            "Failed to set report cell data in %r [%s!%s@%d]",
            report_sheet_path, sheet_name, col_name, row_num
        )
        return False
